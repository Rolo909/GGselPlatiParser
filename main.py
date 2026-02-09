import sys
import requests
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLineEdit, QPushButton, QTableWidget, 
                             QTableWidgetItem, QLabel, QProgressBar, QMessageBox,
                             QHeaderView, QComboBox, QTabWidget, QTextEdit, QScrollArea,
                             QFrame, QGridLayout, QFileDialog, QCheckBox)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QPalette, QColor
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
from collections import defaultdict
from datetime import datetime
import matplotlib
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class AnalyticsEngine:
    """Движок аналитики для расчета всех метрик"""
    
    def __init__(self, products):
        self.products = products
        self.categories = self._group_by_category()
        
    def _group_by_category(self):
        """Группировка товаров по категориям"""
        categories = defaultdict(list)
        for p in self.products:
            cat = p.get('category', 'Без категории')
            if not cat:
                cat = 'Без категории'
            categories[cat].append(p)
        return dict(categories)
    
    def get_category_stats(self):
        """Статистика по каждой категории"""
        stats = []
        
        for cat_name, items in self.categories.items():
            if not items:
                continue
                
            prices = [p['price'] for p in items]
            sales = [p['sales'] for p in items]
            revenues = [p['price'] * p['sales'] for p in items]
            
            avg_price = sum(prices) / len(prices)
            avg_sales = sum(sales) / len(sales)
            total_revenue = sum(revenues)
            competitors = len(items)
            
            # Индекс привлекательности: (средние продажи × средняя цена) / конкуренты
            if competitors > 0:
                attractiveness = (avg_sales * avg_price) / competitors
            else:
                attractiveness = 0
            
            # Определение уровня конкуренции
            if competitors < 20:
                competition_level = "Низкая"
                competition_color = "#27ae60"
            elif competitors < 50:
                competition_level = "Средняя"
                competition_color = "#f39c12"
            else:
                competition_level = "Высокая"
                competition_color = "#e74c3c"
            
            # Определение спроса
            if avg_sales > 500:
                demand_level = "Высокий"
                demand_color = "#27ae60"
            elif avg_sales > 100:
                demand_level = "Средний"
                demand_color = "#f39c12"
            else:
                demand_level = "Низкий"
                demand_color = "#e74c3c"
            
            # Оценка привлекательности (0-10)
            score = 0
            # Спрос (+4 балла)
            if avg_sales > 500:
                score += 4
            elif avg_sales > 100:
                score += 2
            
            # Конкуренция (+3 балла за низкую)
            if competitors < 20:
                score += 3
            elif competitors < 50:
                score += 1.5
            
            # Средний чек (+3 балла)
            if avg_price > 500:
                score += 3
            elif avg_price > 200:
                score += 1.5
            
            # Рекомендация
            if score >= 7:
                recommendation = "ВХОДИТЬ 🚀"
                rec_color = "#27ae60"
            elif score >= 4:
                recommendation = "ИЗУЧИТЬ 🔍"
                rec_color = "#f39c12"
            else:
                recommendation = "НЕ ВХОДИТЬ ⛔"
                rec_color = "#e74c3c"
            
            stats.append({
                'category': cat_name,
                'competitors': competitors,
                'avg_price': avg_price,
                'avg_sales': avg_sales,
                'total_revenue': total_revenue,
                'attractiveness': attractiveness,
                'competition_level': competition_level,
                'competition_color': competition_color,
                'demand_level': demand_level,
                'demand_color': demand_color,
                'score': score,
                'recommendation': recommendation,
                'rec_color': rec_color
            })
        
        # Сортировка по индексу привлекательности
        stats.sort(key=lambda x: x['attractiveness'], reverse=True)
        return stats
    
    def get_price_segments(self):
        """Анализ ценовых сегментов"""
        segments = {
            'Бюджет (<200₽)': [],
            'Средний (200-500₽)': [],
            'Премиум (>500₽)': []
        }
        
        for p in self.products:
            if p['price'] < 200:
                segments['Бюджет (<200₽)'].append(p)
            elif p['price'] <= 500:
                segments['Средний (200-500₽)'].append(p)
            else:
                segments['Премиум (>500₽)'].append(p)
        
        segment_stats = []
        for seg_name, items in segments.items():
            if items:
                avg_sales = sum(p['sales'] for p in items) / len(items)
                total_revenue = sum(p['price'] * p['sales'] for p in items)
                segment_stats.append({
                    'segment': seg_name,
                    'count': len(items),
                    'avg_sales': avg_sales,
                    'total_revenue': total_revenue
                })
        
        return segment_stats
    
    def get_anomalies(self):
        """Поиск трендов и аномалий"""
        anomalies = {
            'premium_demand': [],  # Высокая цена + высокие продажи
            'low_performance': [],  # Низкая цена + низкие продажи
            'opportunities': []  # Цена ниже средней в категории
        }
        
        for cat_name, items in self.categories.items():
            if len(items) < 3:
                continue
            
            avg_price = sum(p['price'] for p in items) / len(items)
            avg_sales = sum(p['sales'] for p in items) / len(items)
            
            for p in items:
                # Премиум-спрос: цена > среднего * 1.5 И продажи > среднего
                if p['price'] > avg_price * 1.5 and p['sales'] > avg_sales:
                    anomalies['premium_demand'].append({
                        **p,
                        'reason': f"Премиум-спрос в '{cat_name}'"
                    })
                
                # Низкая эффективность
                if p['price'] < avg_price * 0.7 and p['sales'] < avg_sales * 0.5:
                    anomalies['low_performance'].append({
                        **p,
                        'reason': f"Низкая эффективность в '{cat_name}'"
                    })
                
                # Возможность: цена ниже средней, но продажи хорошие
                if p['price'] < avg_price * 0.8 and p['sales'] > avg_sales:
                    anomalies['opportunities'].append({
                        **p,
                        'reason': f"Возможность поднять цену в '{cat_name}'",
                        'avg_price': avg_price
                    })
        
        return anomalies
    
    def get_top_products(self, limit=10):
        """ТОП товаров по обороту"""
        products_with_revenue = []
        for p in self.products:
            revenue = p['price'] * p['sales']
            products_with_revenue.append({
                **p,
                'revenue': revenue
            })
        
        products_with_revenue.sort(key=lambda x: x['revenue'], reverse=True)
        return products_with_revenue[:limit]


class ParserThread(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, url, sort_by, product_type):
        super().__init__()
        self.url = url
        self.sort_by = sort_by
        self.product_type = product_type

    def run(self):
        try:
            self.progress.emit(10)
            products = self.parse_page(self.url)
            self.progress.emit(80)
            
            if products:
                if self.sort_by == "Продажи (убывание)":
                    products.sort(key=lambda x: x['sales'], reverse=True)
                elif self.sort_by == "Цена (возрастание)":
                    products.sort(key=lambda x: x['price'])
                elif self.sort_by == "Цена (убывание)":
                    products.sort(key=lambda x: x['price'], reverse=True)
                elif self.sort_by == "Оборот (убывание)":
                    products.sort(key=lambda x: x['price'] * x['sales'], reverse=True)
                
                self.progress.emit(100)
                self.finished.emit(products)
            else:
                self.error.emit("Не удалось извлечь товары с данной страницы")
        except Exception as e:
            self.error.emit(f"Ошибка парсинга: {str(e)}")

    def parse_page(self, url):
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        
        try:
            driver = webdriver.Chrome(options=options)
        except:
            try:
                from webdriver_manager.chrome import ChromeDriverManager
                driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            except:
                self.error.emit("Не удалось инициализировать Chrome WebDriver. Установите ChromeDriver вручную.")
                return []
        
        try:
            driver.get(url)
            time.sleep(3)
            
            is_ggsel = 'ggsel' in url
            is_plati = 'plati' in url
            is_ggsel_main = url.rstrip('/') == 'https://ggsel.net'
            
            # Логика для ggsel.net (главная страница - слайдер)
            if is_ggsel_main:
                try:
                    time.sleep(2)
                    
                    # Сначала прокручиваем слайдер
                    next_buttons = driver.find_elements(By.CSS_SELECTOR, 
                        'button[aria-label="Next slide"], button.swiper-button-next, button[class*="next"]')
                    
                    if next_buttons:
                        max_clicks = 15
                        clicks_done = 0
                        
                        for i in range(max_clicks):
                            try:
                                active_next = None
                                for btn in next_buttons:
                                    try:
                                        if btn.is_displayed() and btn.is_enabled():
                                            classes = btn.get_attribute('class') or ''
                                            if 'disabled' not in classes.lower():
                                                active_next = btn
                                                break
                                    except:
                                        continue
                                
                                if not active_next:
                                    break

                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", active_next)
                                time.sleep(0.3)
                                driver.execute_script("arguments[0].click();", active_next)
                                clicks_done += 1
                                time.sleep(0.5)

                                next_buttons = driver.find_elements(By.CSS_SELECTOR, 
                                    'button[aria-label="Next slide"], button.swiper-button-next, button[class*="next"]')
                                
                            except Exception as e:
                                break
                        
                        time.sleep(2)
                    
                    # ТЕПЕРЬ прокручиваем всю страницу вниз для подгрузки товаров
                    print("Прокрутка главной страницы ggsel.net для загрузки всех товаров...")
                    
                    last_height = driver.execute_script("return document.body.scrollHeight")
                    scroll_attempts = 0
                    max_scroll_attempts = 30
                    no_change_count = 0
                    
                    while scroll_attempts < max_scroll_attempts:
                        # Прокрутка вниз порциями
                        current_position = driver.execute_script("return window.pageYOffset;")
                        scroll_step = 800
                        
                        driver.execute_script(f"window.scrollTo(0, {current_position + scroll_step});")
                        time.sleep(2.5)  # Даем время на автоматическую подгрузку
                        
                        # Проверяем изменение высоты страницы
                        new_height = driver.execute_script("return document.body.scrollHeight")
                        
                        if new_height == last_height:
                            no_change_count += 1
                            # Если высота не менялась 3 раза подряд - достигли конца
                            if no_change_count >= 3:
                                print(f"Достигнут конец главной страницы ({scroll_attempts} прокруток)")
                                break
                                
                            # Еще одна попытка докрутить до самого низа
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                            time.sleep(2)
                        else:
                            no_change_count = 0
                        
                        last_height = new_height
                        scroll_attempts += 1
                        
                        # Дополнительное ожидание каждые 3 прокрутки
                        if scroll_attempts % 3 == 0:
                            time.sleep(1.5)
                    
                    # Финальная прокрутка
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(3)
                    
                    print(f"Прокрутка главной страницы завершена. Выполнено {scroll_attempts} прокруток")
                    
                except Exception as e:
                    print(f"Ошибка при обработке главной страницы: {e}")
            
            # Логика для ggsel.net (категории - кнопка "Показать ещё")
            elif is_ggsel:
                try:
                    print("Загрузка товаров ggsel.net через кнопку 'Показать ещё'...")
                    
                    max_attempts = 30
                    attempts = 0
                    
                    while attempts < max_attempts:
                        try:
                            # Прокрутка вниз
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                            time.sleep(2)
                            
                            # Поиск кнопки по data-test атрибуту (самый надежный селектор)
                            try:
                                button = driver.find_element(By.CSS_SELECTOR, 'button[data-test="showMore"]')
                            except:
                                # Альтернативный поиск по тексту
                                try:
                                    button = driver.find_element(By.XPATH, 
                                        "//button[contains(., 'Показать ещё') or contains(., 'Показать еще')]")
                                except:
                                    print(f"Кнопка 'Показать ещё' не найдена. Все товары загружены ({attempts} нажатий)")
                                    break
                            
                            # Проверка видимости
                            if not button.is_displayed():
                                break
                            
                            # Прокрутка к кнопке
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
                            time.sleep(1)
                            
                            # Клик через JavaScript
                            driver.execute_script("arguments[0].click();", button)
                            attempts += 1
                            print(f"Нажатие на 'Показать ещё' #{attempts}")
                            
                            # Ожидание загрузки
                            time.sleep(3)
                            
                        except Exception as e:
                            print(f"Ошибка при нажатии кнопки: {e}")
                            break
                    
                    # Финальная прокрутка
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(2)
                    
                    print(f"Загрузка завершена. Выполнено {attempts} нажатий на 'Показать ещё'")
                    
                except Exception as e:
                    print(f"Ошибка при загрузке товаров ggsel: {e}")
            
            # Логика для plati.market (динамическая подгрузка при скролле)
            # Логика для plati.market (динамическая подгрузка при скролле)
            elif is_plati:
                try:
                    print("Постепенная прокрутка страницы plati.market...")
                    
                    last_height = driver.execute_script("return document.body.scrollHeight")
                    scroll_attempts = 0
                    max_scroll_attempts = 30  # Увеличено с 20 до 30
                    no_change_count = 0  # Счетчик для отслеживания отсутствия изменений
                    
                    while scroll_attempts < max_scroll_attempts:
                        # Плавная прокрутка по частям
                        current_position = driver.execute_script("return window.pageYOffset;")
                        scroll_step = 500  # Уменьшено с 800 до 500px для более плавной прокрутки
                        
                        driver.execute_script(f"window.scrollTo(0, {current_position + scroll_step});")
                        time.sleep(2)  # Увеличено время ожидания для загрузки
                        
                        # Проверка достижения конца страницы
                        new_height = driver.execute_script("return document.body.scrollHeight")
                        
                        if new_height == last_height:
                            no_change_count += 1
                            # Если высота не менялась 3 раза подряд - достигли конца
                            if no_change_count >= 3:
                                print(f"Достигнут конец страницы plati.market ({scroll_attempts} прокруток)")
                                break
                                
                            # Попытка прокрутки в самый низ
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                            time.sleep(2)
                        else:
                            no_change_count = 0  # Сброс счетчика при изменении высоты
                        
                        last_height = new_height
                        scroll_attempts += 1
                        
                        # Дополнительное ожидание каждые 3 прокрутки
                        if scroll_attempts % 3 == 0:
                            time.sleep(1.5)
                    
                    # Финальная прокрутка вниз
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(3)
                    
                    print(f"Прокрутка завершена. Выполнено {scroll_attempts} прокруток")
                    
                except Exception as e:
                    print(f"Ошибка при прокрутке plati: {e}")
            
            # Для остальных сайтов - стандартная прокрутка
            else:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2);")
                time.sleep(1)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
            
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            
            if is_ggsel:
                return self.parse_ggsel(soup, url)
            elif is_plati:
                return self.parse_plati(soup, url)
            else:
                return self.parse_generic(soup, url)
        finally:
            driver.quit()

    def parse_ggsel(self, soup, base_url):
        products = []
        
        # Исключаем блок "Рекомендовано вам" по классу BottomGoods_cards__5r9XZ
        bottom_goods = soup.find('div', class_='BottomGoods_cards__5r9XZ')
        if bottom_goods:
            bottom_goods.decompose()
        
        # Теперь ищем все карточки товаров
        items = soup.find_all('div', {'class': 'ProductCard_card__zjTV_'})
        
        print(f"Найдено карточек товаров на ggsel: {len(items)}")
        
        for item in items[:500]:
            try:
                category_elem = item.find('div', {'data-testid': 'card-category'})
                category = category_elem.get_text(strip=True) if category_elem else ''

                if self.product_type != "Все":
                    if category != self.product_type:
                        continue

                name_elem = item.find('span', {'class': 'ProductCard_description__AXXxp'})
                if not name_elem:
                    name_elem = item.find('div', {'data-testid': 'card-description'})

                price_elem = item.find('div', {'data-testid': 'card-price'})
                if not price_elem:
                    price_elem = item.find('span', {'class': 'ProductCard_price__k1Ahq'})

                sales_elem = item.find('div', {'data-testid': 'card-counter'})
                link_elem = item.find('a', {'data-testid': 'card-link'})
                if not link_elem:
                    link_elem = item.find('a', {'data-testid': 'card-button'})
                
                if name_elem and price_elem:
                    title = name_elem.get_text(strip=True)
                    price_text = price_elem.get_text(strip=True)

                    if '=' in price_text:
                        price_text = price_text.split('=')[-1].strip()
                    
                    price_clean = price_text.replace('₽', '').replace('\xa0', '').replace(' ', '').strip()
                    
                    try:
                        price = float(price_clean)
                    except:
                        continue
                    
                    sales = 0
                    if sales_elem:
                        sales_text = sales_elem.get_text(strip=True)
                        sales_text = sales_text.replace('\xa0', '').replace(' ', '')
                        
                        if '+' in sales_text:
                            sales_match = re.search(r'(\d+)\+', sales_text)
                            if sales_match:
                                sales = int(sales_match.group(1))
                        else:
                            sales_match = re.search(r'(\d+)', sales_text)
                            if sales_match:
                                sales = int(sales_match.group(1))
                    
                    link = ''
                    if link_elem and link_elem.get('href'):
                        link = link_elem['href']
                        if not link.startswith('http'):
                            link = 'https://ggsel.net' + link
                    
                    products.append({
                        'name': title,
                        'price': price,
                        'sales': sales,
                        'link': link,
                        'category': category
                    })
            except Exception as e:
                continue
        
        print(f"Успешно распарсено товаров ggsel: {len(products)}")
        return products

    def parse_plati(self, soup, base_url):
        products = []

        # Удаляем блоки, которые не нужно парсить:
        # 1. "Рекомендуем" (id="rec_wrapper")
        # 2. "Вы смотрели" (id="hist_wrapper")
        # 3. "Лучшее предложение" (class="suggestion" или "best-offer-slider")
        
        blocks_to_remove = [
            soup.find('div', id='rec_wrapper'),
            soup.find('div', id='hist_wrapper'),
            soup.find('section', class_='suggestion'),
            soup.find('div', class_='best-offer-slider')
        ]
        
        for block in blocks_to_remove:
            if block:
                block.decompose()
        
        # Теперь ищем все карточки товаров
        cards = soup.find_all('a', class_='card')
        
        print(f"Найдено карточек товаров на plati (после фильтрации): {len(cards)}")
        
        for card in cards[:300]:
            try:
                if 'd-none' in card.parent.get('class', []):
                    continue

                title_span = card.find('span', class_='footnote-medium')
                if not title_span:
                    title_p = card.find('p', class_='custom-link')
                    if title_p:
                        title_span = title_p.find('span', class_='footnote-medium')
                title = title_span.get_text(strip=True) if title_span else ''
                if not title:
                    continue

                price_span = card.find('span', class_='title-bold')
                if not price_span:
                    continue
                price_text = price_span.get_text(strip=True)
                price_clean = re.sub(r'[^\d.]', '', price_text.replace(',', '.'))
                try:
                    price = float(price_clean)
                except:
                    continue
                
                sales = 0
                sold_span = card.find('span', class_='footnote-regular')
                if sold_span:
                    sold_text = sold_span.get_text(strip=True)
                    sold_text = sold_text.replace('\xa0', '').replace(' ', '')
                    
                    if 'млн' in sold_text:
                        m = re.search(r'(\d+(?:\.\d+)?)', sold_text)
                        if m:
                            sales = int(float(m.group(1)) * 1_000_000)
                    elif 'тыс' in sold_text or 'k' in sold_text.lower():
                        m = re.search(r'(\d+(?:\.\d+)?)', sold_text)
                        if m:
                            sales = int(float(m.group(1)) * 1_000)
                    elif '+' in sold_text:
                        m = re.search(r'(\d+)', sold_text)
                        if m:
                            sales = int(m.group(1))
                    else:
                        m = re.search(r'(\d+)', sold_text)
                        if m:
                            sales = int(m.group(1))
                    if 'менее' in sold_text.lower():
                        sales = 5

                link = card.get('href', '')
                if link and not link.startswith('http'):
                    link = 'https://plati.market' + link
                
                products.append({
                    'name': title,
                    'price': price,
                    'sales': sales,
                    'link': link,
                    'category': ''
                })
                
            except Exception as e:
                continue
        
        print(f"Успешно распарсено товаров plati: {len(products)}")
        return products

    def parse_generic(self, soup, base_url):
        products = []
        patterns = [
            {'container': re.compile(r'product|item|card|goods|offer'), 
             'title': re.compile(r'title|name|product|heading'),
             'price': re.compile(r'price|cost|sum|rub'),
             'sales': re.compile(r'sold|sales|продаж|купили')},
        ]
        
        for pattern in patterns:
            items = soup.find_all(['div', 'article', 'li', 'section'], class_=pattern['container'])
            
            for item in items[:50]:
                try:
                    title_elem = item.find(['h1', 'h2', 'h3', 'h4', 'a', 'span', 'div'], 
                                          class_=pattern['title'])
                    price_elem = item.find(['span', 'div', 'p', 'strong'], 
                                          class_=pattern['price'])
                    sales_elem = item.find(['span', 'div'], 
                                          text=re.compile(r'\d+\s*(продаж|sold|sales)', re.I))
                    link_elem = item.find('a', href=True)
                    
                    if title_elem and price_elem:
                        title = title_elem.get_text(strip=True)
                        price_text = price_elem.get_text(strip=True)
                        price_match = re.search(r'[\d\s]+[.,]?\d*', price_text)
                        if price_match:
                            price = float(re.sub(r'[^\d.]', '', price_match.group().replace(',', '.')))
                        else:
                            continue
                        
                        sales = 0
                        if sales_elem:
                            sales_text = sales_elem.get_text(strip=True)
                            sales_match = re.search(r'(\d+)', sales_text)
                            if sales_match:
                                sales = int(sales_match.group(1))
                        
                        link = link_elem['href'] if link_elem else ''
                        if link and not link.startswith('http'):
                            if link.startswith('/'):
                                link = base_url.split('/')[0] + '//' + base_url.split('/')[2] + link
                            else:
                                link = base_url.rsplit('/', 1)[0] + '/' + link
                        
                        products.append({
                            'name': title[:100],
                            'price': price,
                            'sales': sales,
                            'link': link,
                            'category': ''
                        })
                except:
                    continue
            
            if products:
                break
        
        return products


class ChartWidget(QWidget):
    """Виджет для отображения графиков"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.figure = Figure(figsize=(12, 8))
        self.canvas = FigureCanvas(self.figure)
        
        layout = QVBoxLayout(self)
        layout.addWidget(self.canvas)
    
    def plot_price_vs_sales(self, products):
        """График цена vs продажи"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        # Группировка по категориям для цветов
        categories = {}
        for p in products:
            cat = p.get('category', 'Без категории')
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(p)
        
        colors = plt.cm.tab10(range(len(categories)))
        
        for idx, (cat_name, items) in enumerate(categories.items()):
            prices = [p['price'] for p in items]
            sales = [p['sales'] for p in items]
            ax.scatter(prices, sales, label=cat_name, alpha=0.6, s=100, color=colors[idx])
        
        ax.set_xlabel('Цена (₽)', fontsize=12)
        ax.set_ylabel('Продажи', fontsize=12)
        ax.set_title('Распределение: Цена vs Продажи', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    def plot_category_pie(self, category_stats):
        """Круговая диаграмма категорий"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        labels = [s['category'] for s in category_stats]
        revenues = [s['total_revenue'] for s in category_stats]
        
        ax.pie(revenues, labels=labels, autopct='%1.1f%%', startangle=90)
        ax.set_title('Доли категорий по обороту', fontsize=14, fontweight='bold')
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    def plot_top_niches(self, category_stats):
        """ТОП-10 ниш"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        top_10 = category_stats[:10]
        categories = [s['category'][:20] for s in top_10]
        attractiveness = [s['attractiveness'] for s in top_10]
        
        bars = ax.barh(categories, attractiveness, color='#667eea')
        ax.set_xlabel('Индекс привлекательности', fontsize=12)
        ax.set_title('ТОП-10 самых привлекательных ниш', fontsize=14, fontweight='bold')
        ax.invert_yaxis()
        
        # Добавление значений на бары
        for bar in bars:
            width = bar.get_width()
            ax.text(width, bar.get_y() + bar.get_height()/2, 
                   f'{width:.0f}', ha='left', va='center', fontsize=9)
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    def plot_price_segments(self, segment_stats):
        """Анализ ценовых сегментов"""
        self.figure.clear()
        
        # Два подграфика
        ax1 = self.figure.add_subplot(121)
        ax2 = self.figure.add_subplot(122)
        
        segments = [s['segment'] for s in segment_stats]
        counts = [s['count'] for s in segment_stats]
        avg_sales = [s['avg_sales'] for s in segment_stats]
        
        # График 1: Количество товаров
        ax1.bar(segments, counts, color=['#3498db', '#f39c12', '#9b59b6'])
        ax1.set_ylabel('Количество товаров', fontsize=11)
        ax1.set_title('Товары по сегментам', fontsize=12, fontweight='bold')
        ax1.tick_params(axis='x', rotation=15)
        
        # График 2: Средние продажи
        ax2.bar(segments, avg_sales, color=['#3498db', '#f39c12', '#9b59b6'])
        ax2.set_ylabel('Средние продажи', fontsize=11)
        ax2.set_title('Эффективность сегментов', fontsize=12, fontweight='bold')
        ax2.tick_params(axis='x', rotation=15)
        
        self.figure.tight_layout()
        self.canvas.draw()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Анализатор популярности товаров")
        self.setGeometry(50, 50, 1800, 1000)
        
        self.products = []
        self.analytics = None
        
        self.setup_ui()
        
    def setup_ui(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)

        main_widget.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8f9fa, stop:1 #e9ecef);
            }
        """)
        
        layout = QVBoxLayout(main_widget)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Заголовок
        title_container = QWidget()
        title_container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #667eea, stop:1 #764ba2);
                border-radius: 12px;
                padding: 20px;
            }
        """)
        title_layout = QVBoxLayout(title_container)
        
        title = QLabel("📊 Анализатор популярности товаров")
        title.setFont(QFont("Segoe UI", 24, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: white; background: transparent;")
        title_layout.addWidget(title)
        
        layout.addWidget(title_container)

        # URL
        url_container = QWidget()
        url_container.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 12px;
                padding: 15px;
            }
        """)
        url_layout = QHBoxLayout(url_container)
        url_layout.setSpacing(10)
        
        url_label = QLabel("🔗 URL:")
        url_label.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        url_label.setStyleSheet("color: #495057; background: transparent;")
        url_layout.addWidget(url_label)
        
        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText("Введите ссылку на страницу с товарами (например: https://ggsel.net)")
        self.url_input.setFont(QFont("Segoe UI", 11))
        self.url_input.setStyleSheet("""
            QLineEdit {
                padding: 12px 15px;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                background-color: #f8f9fa;
                color: #212529;
            }
            QLineEdit:focus {
                border: 2px solid #667eea;
                background-color: white;
            }
        """)
        self.url_input.setMinimumHeight(45)
        url_layout.addWidget(self.url_input)
        
        layout.addWidget(url_container)

        # Контролы
        controls_container = QWidget()
        controls_container.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 12px;
                padding: 15px;
            }
        """)
        controls_layout = QHBoxLayout(controls_container)
        controls_layout.setSpacing(15)
 
        sort_label = QLabel("📋 Сортировка:")
        sort_label.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        sort_label.setStyleSheet("color: #495057; background: transparent;")
        controls_layout.addWidget(sort_label)
        
        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["Продажи (убывание)", "Оборот (убывание)", "Цена (возрастание)", "Цена (убывание)"])
        self.sort_combo.setFont(QFont("Segoe UI", 11))
        self.sort_combo.setStyleSheet(self._get_combo_style())
        self.sort_combo.setMinimumHeight(45)
        controls_layout.addWidget(self.sort_combo)

        type_label = QLabel("🏷️ Тип:")
        type_label.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        type_label.setStyleSheet("color: #495057; background: transparent;")
        controls_layout.addWidget(type_label)
        
        self.type_combo = QComboBox()
        self.type_combo.addItems(["Все", "Ключ", "Гифт", "DLC", "Пополнение"])
        self.type_combo.setFont(QFont("Segoe UI", 11))
        self.type_combo.setStyleSheet(self._get_combo_style())
        self.type_combo.setMinimumHeight(45)
        controls_layout.addWidget(self.type_combo)
        
        controls_layout.addStretch()

        self.parse_button = QPushButton("🚀 НАЧАТЬ АНАЛИЗ")
        self.parse_button.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        self.parse_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #667eea, stop:1 #764ba2);
                color: white;
                padding: 12px 30px;
                border: none;
                border-radius: 10px;
                min-width: 200px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5568d3, stop:1 #6a3f8f);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4a5ab8, stop:1 #5d3679);
            }
            QPushButton:disabled {
                background: #adb5bd;
            }
        """)
        self.parse_button.setMinimumHeight(50)
        self.parse_button.clicked.connect(self.start_parsing)
        controls_layout.addWidget(self.parse_button)
        
        # Кнопка экспорта
        self.export_button = QPushButton("📥 Экспорт в Excel")
        self.export_button.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        self.export_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #229954);
                color: white;
                padding: 12px 25px;
                border: none;
                border-radius: 10px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #1e8449);
            }
            QPushButton:disabled {
                background: #adb5bd;
            }
        """)
        self.export_button.setMinimumHeight(50)
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setEnabled(False)
        controls_layout.addWidget(self.export_button)
        
        layout.addWidget(controls_container)

        # Прогресс-бар
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: none;
                border-radius: 8px;
                text-align: center;
                background-color: #e9ecef;
                color: #495057;
                font-weight: bold;
                font-size: 12px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #667eea, stop:1 #764ba2);
                border-radius: 8px;
            }
        """)
        self.progress_bar.setMinimumHeight(30)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)

        self.status_label = QLabel("✅ Готов к работе")
        self.status_label.setFont(QFont("Segoe UI", 11))
        self.status_label.setStyleSheet("color: #495057; padding: 5px; background: transparent;")
        layout.addWidget(self.status_label)

        # Вкладки
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background: white;
                border-radius: 12px;
            }
            QTabBar::tab {
                background: #e9ecef;
                color: #495057;
                padding: 12px 24px;
                margin-right: 4px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-size: 13px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background: white;
                color: #667eea;
            }
            QTabBar::tab:hover {
                background: #dee2e6;
            }
        """)
        
        # Вкладка 1: Таблица товаров
        self.products_tab = QWidget()
        products_layout = QVBoxLayout(self.products_tab)
        
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Название товара", "Категория", "Цена (₽)", "Продажи", "Оборот (₽)", "Ссылка"])
        self.table.setFont(QFont("Segoe UI", 10))
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        
        self.table.setStyleSheet("""
            QTableWidget {
                border: none;
                border-radius: 12px;
                background-color: white;
                gridline-color: #e9ecef;
            }
            QTableWidget::item {
                padding: 10px;
                color: #212529;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
                color: white;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #495057, stop:1 #343a40);
                color: white;
                padding: 12px;
                border: none;
                font-weight: bold;
            }
        """)
        
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        products_layout.addWidget(self.table)
        
        self.results_label = QLabel("📦 Результатов: 0")
        self.results_label.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        self.results_label.setStyleSheet("color: #495057; padding: 10px; background: transparent;")
        products_layout.addWidget(self.results_label)
        
        self.tabs.addTab(self.products_tab, "📦 Все товары")
        
        # Вкладка 2: Анализ категорий
        self.analytics_tab = QWidget()
        analytics_layout = QVBoxLayout(self.analytics_tab)
        
        self.analytics_table = QTableWidget()
        self.analytics_table.setColumnCount(8)
        self.analytics_table.setHorizontalHeaderLabels([
            "Категория", "Конкурентов", "Ср. цена (₽)", "Ср. продажи", 
            "Оборот (₽)", "Спрос", "Конкуренция", "Рекомендация"
        ])
        self.analytics_table.setFont(QFont("Segoe UI", 10))
        self.analytics_table.setStyleSheet(self.table.styleSheet())
        
        ah = self.analytics_table.horizontalHeader()
        ah.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for i in range(1, 8):
            ah.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        
        self.analytics_table.verticalHeader().setVisible(False)
        self.analytics_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        analytics_layout.addWidget(self.analytics_table)
        self.tabs.addTab(self.analytics_tab, "📊 Анализ ниш")
        
        # Вкладка 3: Возможности
        self.opportunities_tab = QWidget()
        opp_layout = QVBoxLayout(self.opportunities_tab)
        
        self.opportunities_text = QTextEdit()
        self.opportunities_text.setReadOnly(True)
        self.opportunities_text.setFont(QFont("Segoe UI", 11))
        self.opportunities_text.setStyleSheet("""
            QTextEdit {
                background: white;
                border: none;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        
        opp_layout.addWidget(self.opportunities_text)
        self.tabs.addTab(self.opportunities_tab, "💡 Возможности")
        
        # Вкладка 4: Графики
        self.charts_tab = QWidget()
        charts_layout = QVBoxLayout(self.charts_tab)
        
        chart_buttons = QHBoxLayout()
        
        btn1 = QPushButton("📈 Цена vs Продажи")
        btn1.clicked.connect(lambda: self.show_chart('price_sales'))
        btn1.setStyleSheet(self._get_button_style())
        chart_buttons.addWidget(btn1)
        
        btn2 = QPushButton("🥧 Доли категорий")
        btn2.clicked.connect(lambda: self.show_chart('pie'))
        btn2.setStyleSheet(self._get_button_style())
        chart_buttons.addWidget(btn2)
        
        btn3 = QPushButton("🏆 ТОП-10 ниш")
        btn3.clicked.connect(lambda: self.show_chart('top_niches'))
        btn3.setStyleSheet(self._get_button_style())
        chart_buttons.addWidget(btn3)
        
        btn4 = QPushButton("💰 Ценовые сегменты")
        btn4.clicked.connect(lambda: self.show_chart('segments'))
        btn4.setStyleSheet(self._get_button_style())
        chart_buttons.addWidget(btn4)
        
        charts_layout.addLayout(chart_buttons)
        
        self.chart_widget = ChartWidget()
        charts_layout.addWidget(self.chart_widget)
        
        self.tabs.addTab(self.charts_tab, "📊 Графики")
        
        layout.addWidget(self.tabs)

    def _get_combo_style(self):
        return """
            QComboBox {
                padding: 10px 15px;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                background-color: #f8f9fa;
                color: #212529;
                min-width: 150px;
            }
            QComboBox:hover {
                border: 2px solid #667eea;
            }
            QComboBox::drop-down {
                border: none;
                padding-right: 10px;
            }
            QComboBox QAbstractItemView {
                border: 2px solid #dee2e6;
                border-radius: 8px;
                background-color: white;
                selection-background-color: #667eea;
                selection-color: white;
            }
        """
    
    def _get_button_style(self):
        return """
            QPushButton {
                background: #667eea;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 8px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #5568d3;
            }
        """

    def start_parsing(self):
        url = self.url_input.text().strip()
        
        if not url:
            QMessageBox.warning(self, "Ошибка", "Введите URL страницы")
            return
        
        if not url.startswith('http'):
            QMessageBox.warning(self, "Ошибка", "URL должен начинаться с http:// или https://")
            return
        
        self.parse_button.setEnabled(False)
        self.export_button.setEnabled(False)
        self.table.setRowCount(0)
        self.analytics_table.setRowCount(0)
        self.opportunities_text.clear()
        self.progress_bar.setValue(0)
        self.status_label.setText("⏳ Загрузка страницы и анализ товаров...")
        self.status_label.setStyleSheet("color: #667eea; font-weight: bold; background: transparent;")
        
        self.parser_thread = ParserThread(url, self.sort_combo.currentText(), self.type_combo.currentText())
        self.parser_thread.progress.connect(self.update_progress)
        self.parser_thread.finished.connect(self.show_results)
        self.parser_thread.error.connect(self.show_error)
        self.parser_thread.start()

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def show_results(self, products):
        self.products = products
        self.analytics = AnalyticsEngine(products)
        
        # Заполнение таблицы товаров
        self.table.setRowCount(len(products))
        
        for i, product in enumerate(products):
            revenue = product['price'] * product['sales']
            
            name_item = QTableWidgetItem(product['name'])
            name_item.setFont(QFont("Arial", 11))
            self.table.setItem(i, 0, name_item)
            
            category_item = QTableWidgetItem(product.get('category', ''))
            category_item.setFont(QFont("Arial", 11))
            category_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(i, 1, category_item)
            
            price_item = QTableWidgetItem(f"{product['price']:.2f}")
            price_item.setFont(QFont("Arial", 11))
            price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(i, 2, price_item)
            
            sales_item = QTableWidgetItem(str(product['sales']))
            sales_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            sales_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if product['sales'] > 100:
                sales_item.setForeground(QColor("#27ae60"))
            elif product['sales'] > 50:
                sales_item.setForeground(QColor("#f39c12"))
            else:
                sales_item.setForeground(QColor("#95a5a6"))
            self.table.setItem(i, 3, sales_item)
            
            revenue_item = QTableWidgetItem(f"{revenue:,.0f}")
            revenue_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            revenue_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            revenue_item.setForeground(QColor("#667eea"))
            self.table.setItem(i, 4, revenue_item)
            
            link_item = QTableWidgetItem(product['link'])
            link_item.setFont(QFont("Arial", 10))
            link_item.setForeground(QColor("#3498db"))
            self.table.setItem(i, 5, link_item)
        
        # Заполнение аналитики
        self.fill_analytics()
        
        # Заполнение возможностей
        self.fill_opportunities()
        
        self.results_label.setText(f"📦 Результатов: {len(products)}")
        self.status_label.setText("✅ Анализ завершен успешно!")
        self.status_label.setStyleSheet("color: #28a745; font-weight: bold; background: transparent;")
        self.progress_bar.setValue(100)
        self.parse_button.setEnabled(True)
        self.export_button.setEnabled(True)
    
    def fill_analytics(self):
        """Заполнение таблицы аналитики"""
        category_stats = self.analytics.get_category_stats()
        
        self.analytics_table.setRowCount(len(category_stats))
        
        for i, stat in enumerate(category_stats):
            # Категория
            cat_item = QTableWidgetItem(stat['category'])
            cat_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            self.analytics_table.setItem(i, 0, cat_item)
            
            # Конкуренты
            comp_item = QTableWidgetItem(str(stat['competitors']))
            comp_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.analytics_table.setItem(i, 1, comp_item)
            
            # Средняя цена
            price_item = QTableWidgetItem(f"{stat['avg_price']:.2f}")
            price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.analytics_table.setItem(i, 2, price_item)
            
            # Средние продажи
            sales_item = QTableWidgetItem(f"{stat['avg_sales']:.0f}")
            sales_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.analytics_table.setItem(i, 3, sales_item)
            
            # Оборот
            revenue_item = QTableWidgetItem(f"{stat['total_revenue']:,.0f}")
            revenue_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            revenue_item.setForeground(QColor("#667eea"))
            revenue_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            self.analytics_table.setItem(i, 4, revenue_item)
            
            # Спрос
            demand_item = QTableWidgetItem(stat['demand_level'])
            demand_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            demand_item.setForeground(QColor(stat['demand_color']))
            demand_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            self.analytics_table.setItem(i, 5, demand_item)
            
            # Конкуренция
            comp_level_item = QTableWidgetItem(stat['competition_level'])
            comp_level_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            comp_level_item.setForeground(QColor(stat['competition_color']))
            comp_level_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            self.analytics_table.setItem(i, 6, comp_level_item)
            
            # Рекомендация
            rec_item = QTableWidgetItem(stat['recommendation'])
            rec_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            rec_item.setForeground(QColor(stat['rec_color']))
            rec_item.setFont(QFont("Arial", 12, QFont.Weight.Bold))
            self.analytics_table.setItem(i, 7, rec_item)
    
    def fill_opportunities(self):
        """Заполнение возможностей"""
        anomalies = self.analytics.get_anomalies()
        segment_stats = self.analytics.get_price_segments()
        top_products = self.analytics.get_top_products()
        
        html = "<html><body style='font-family: Segoe UI; padding: 20px;'>"
        
        # ТОП-10 товаров
        html += "<h2 style='color: #667eea;'>🏆 ТОП-10 товаров по обороту</h2>"
        html += "<table style='width: 100%; border-collapse: collapse;'>"
        html += "<tr style='background: #f8f9fa; font-weight: bold;'>"
        html += "<th style='padding: 10px; text-align: left;'>Товар</th>"
        html += "<th style='padding: 10px;'>Цена</th>"
        html += "<th style='padding: 10px;'>Продажи</th>"
        html += "<th style='padding: 10px;'>Оборот</th></tr>"
        
        for p in top_products:
            html += f"<tr style='border-bottom: 1px solid #e9ecef;'>"
            html += f"<td style='padding: 8px;'>{p['name'][:60]}</td>"
            html += f"<td style='padding: 8px; text-align: center;'>{p['price']:.2f} ₽</td>"
            html += f"<td style='padding: 8px; text-align: center;'>{p['sales']}</td>"
            html += f"<td style='padding: 8px; text-align: center; color: #667eea; font-weight: bold;'>{p['revenue']:,.0f} ₽</td>"
            html += "</tr>"
        
        html += "</table><br><br>"
        
        # Премиум-спрос
        if anomalies['premium_demand']:
            html += "<h2 style='color: #27ae60;'>💎 Премиум-спрос (высокая цена + высокие продажи)</h2>"
            html += "<ul>"
            for item in anomalies['premium_demand'][:5]:
                html += f"<li><b>{item['name'][:60]}</b> - {item['price']:.2f} ₽, продажи: {item['sales']}<br>"
                html += f"<i style='color: #666;'>{item['reason']}</i></li>"
            html += "</ul><br>"
        
        # Возможности
        if anomalies['opportunities']:
            html += "<h2 style='color: #f39c12;'>💡 Возможности (можно поднять цену)</h2>"
            html += "<ul>"
            for item in anomalies['opportunities'][:5]:
                potential_gain = (item['avg_price'] - item['price']) * item['sales']
                html += f"<li><b>{item['name'][:60]}</b><br>"
                html += f"Текущая цена: {item['price']:.2f} ₽ | Средняя в категории: {item['avg_price']:.2f} ₽<br>"
                html += f"<span style='color: #27ae60; font-weight: bold;'>Потенциал +{potential_gain:,.0f} ₽</span><br>"
                html += f"<i style='color: #666;'>{item['reason']}</i></li>"
            html += "</ul><br>"
        
        # Низкая эффективность
        if anomalies['low_performance']:
            html += "<h2 style='color: #e74c3c;'>⚠️ Низкая эффективность (требуется улучшение)</h2>"
            html += "<ul>"
            for item in anomalies['low_performance'][:5]:
                html += f"<li><b>{item['name'][:60]}</b> - {item['price']:.2f} ₽, продажи: {item['sales']}<br>"
                html += f"<i style='color: #666;'>{item['reason']}</i></li>"
            html += "</ul><br>"
        
        # Анализ сегментов
        html += "<h2 style='color: #667eea;'>💰 Анализ ценовых сегментов</h2>"
        html += "<table style='width: 100%; border-collapse: collapse;'>"
        html += "<tr style='background: #f8f9fa; font-weight: bold;'>"
        html += "<th style='padding: 10px;'>Сегмент</th>"
        html += "<th style='padding: 10px;'>Товаров</th>"
        html += "<th style='padding: 10px;'>Ср. продажи</th>"
        html += "<th style='padding: 10px;'>Оборот</th></tr>"
        
        for seg in segment_stats:
            html += f"<tr style='border-bottom: 1px solid #e9ecef;'>"
            html += f"<td style='padding: 8px; text-align: center;'>{seg['segment']}</td>"
            html += f"<td style='padding: 8px; text-align: center;'>{seg['count']}</td>"
            html += f"<td style='padding: 8px; text-align: center;'>{seg['avg_sales']:.0f}</td>"
            html += f"<td style='padding: 8px; text-align: center; font-weight: bold;'>{seg['total_revenue']:,.0f} ₽</td>"
            html += "</tr>"
        
        html += "</table>"
        
        html += "</body></html>"
        
        self.opportunities_text.setHtml(html)
    
    def show_chart(self, chart_type):
        """Показать график"""
        if not self.analytics:
            QMessageBox.warning(self, "Ошибка", "Сначала выполните анализ")
            return
        
        try:
            if chart_type == 'price_sales':
                self.chart_widget.plot_price_vs_sales(self.products)
            elif chart_type == 'pie':
                category_stats = self.analytics.get_category_stats()
                self.chart_widget.plot_category_pie(category_stats)
            elif chart_type == 'top_niches':
                category_stats = self.analytics.get_category_stats()
                self.chart_widget.plot_top_niches(category_stats)
            elif chart_type == 'segments':
                segment_stats = self.analytics.get_price_segments()
                self.chart_widget.plot_price_segments(segment_stats)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось построить график: {str(e)}")

    def export_to_excel(self):
        """Экспорт отчета в Excel"""
        if not self.products:
            QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчет", 
            f"market_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            wb = Workbook()
            
            # Лист 1: Все товары
            ws1 = wb.active
            ws1.title = "Товары"
            
            headers = ["Название", "Категория", "Цена", "Продажи", "Оборот", "Ссылка"]
            ws1.append(headers)
            
            for col in range(1, len(headers) + 1):
                ws1.cell(1, col).font = Font(bold=True, color="FFFFFF")
                ws1.cell(1, col).fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                ws1.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
            
            for p in self.products:
                ws1.append([
                    p['name'],
                    p.get('category', ''),
                    p['price'],
                    p['sales'],
                    p['price'] * p['sales'],
                    p['link']
                ])
            
            # Автоширина
            for col in ws1.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws1.column_dimensions[column].width = adjusted_width
            
            # Лист 2: Анализ ниш
            ws2 = wb.create_sheet("Анализ ниш")
            category_stats = self.analytics.get_category_stats()
            
            headers2 = ["Категория", "Конкуренты", "Ср. цена", "Ср. продажи", "Оборот", "Индекс", "Рекомендация"]
            ws2.append(headers2)
            
            for col in range(1, len(headers2) + 1):
                ws2.cell(1, col).font = Font(bold=True, color="FFFFFF")
                ws2.cell(1, col).fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                ws2.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
            
            for stat in category_stats:
                ws2.append([
                    stat['category'],
                    stat['competitors'],
                    round(stat['avg_price'], 2),
                    round(stat['avg_sales'], 0),
                    round(stat['total_revenue'], 0),
                    round(stat['attractiveness'], 2),
                    stat['recommendation']
                ])
            
            for col in ws2.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                ws2.column_dimensions[column].width = adjusted_width
            
            # Лист 3: ТОП товаров
            ws3 = wb.create_sheet("ТОП-10")
            top_products = self.analytics.get_top_products()
            
            headers3 = ["Название", "Категория", "Цена", "Продажи", "Оборот"]
            ws3.append(headers3)
            
            for col in range(1, len(headers3) + 1):
                ws3.cell(1, col).font = Font(bold=True, color="FFFFFF")
                ws3.cell(1, col).fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
                ws3.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
            
            for p in top_products:
                ws3.append([
                    p['name'],
                    p.get('category', ''),
                    p['price'],
                    p['sales'],
                    p['revenue']
                ])
            
            for col in ws3.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws3.column_dimensions[column].width = adjusted_width
            
            wb.save(filename)
            
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать: {str(e)}")

    def show_error(self, error_msg):
        self.status_label.setText(f"❌ Ошибка: {error_msg}")
        self.status_label.setStyleSheet("color: #dc3545; font-weight: bold; background: transparent;")
        self.progress_bar.setValue(0)
        self.parse_button.setEnabled(True)
        QMessageBox.critical(self, "Ошибка", error_msg)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(240, 240, 245))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(44, 62, 80))
    app.setPalette(palette)
    
    window = MainWindow()
    window.show()
    sys.exit(app.exec())